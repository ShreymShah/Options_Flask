from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import pandas as pd
import yfinance as yf
import openpyxl
import threading
import time as t
import os
from pya3 import *

app = Flask(__name__, template_folder='template')
app.secret_key = "DONT"

# Path for the backtest files
if not os.path.exists('downloads'):
    os.makedirs('downloads')

def backtest_to_excel(symbol: str, backtest_years: int, timeframe: str, filename: str):
    if timeframe not in ['1d', '1wk', '1mo']:
        raise ValueError("Timeframe must be one of '1d', '1wk', or '1mo'")

    data = yf.download(symbol, period=f'{backtest_years}y', interval=timeframe)
    data.index = data.index.tz_localize(None)
    data['Date'] = data.index.strftime('%d/%m/%Y')

    data['High-Open %'] = ((data['High'] - data['Open']) / data['Open']) * 100
    data['Low-Open %'] = ((data['Low'] - data['Open']) / data['Open']) * 100

    data['High-Open %'] = data['High-Open %'].round(2)
    data['Low-Open %'] = data['Low-Open %'].round(2)

    high_open_sorted = data[['Date', 'High-Open %']].sort_values(by='High-Open %', ascending=False)
    low_open_sorted = data[['Date', 'Low-Open %']].sort_values(by='Low-Open %', ascending=True)

    high_open_sorted.rename(columns={'Date': 'Date (High-Open %)'}, inplace=True)
    low_open_sorted.rename(columns={'Date': 'Date (Low-Open %)'}, inplace=True)

    high_open_sorted.reset_index(drop=True, inplace=True)
    low_open_sorted.reset_index(drop=True, inplace=True)

    blank_col = pd.DataFrame({"": [""] * len(high_open_sorted)})

    combined_data = pd.concat([high_open_sorted, blank_col, low_open_sorted], axis=1)

    combined_data.columns = [
        "Date (High-Open %)", "High-Open %", "",
        "Date (Low-Open %)", "Low-Open %"
    ]

    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('Backtest Data')
        writer.sheets['Backtest Data'] = worksheet

        worksheet.write(0, 0, f"Symbol: {symbol}")
        worksheet.write(1, 0, f"Backtest Years: {backtest_years}")
        worksheet.write(2, 0, f"Timeframe: {timeframe}")

        combined_data.to_excel(writer, sheet_name='Backtest Data', startrow=4, index=False)

        for i, col in enumerate(combined_data.columns):
            max_len = max(
                combined_data[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            worksheet.set_column(i, i, max_len)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/add_account', methods=["POST", "GET"])
def add_account():
    if request.method == "POST":
        username = request.form["username"]
        api_key = request.form["api_key"]
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        sheet.cell(row=sheet.max_row + 1, column=1).value = username
        sheet.cell(row=sheet.max_row, column=2).value = api_key
        wb.save("data.xlsx")
        flash("Account added")
        return redirect(url_for("add_account"))
    else:
        return render_template("add_account.html")

@app.route("/accounts")
def accounts():
    wb = openpyxl.load_workbook("data.xlsx")
    sheet = wb.active
    usernames = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]
    return render_template("accounts.html", username=usernames)

@app.route("/delete_account/<username>")
def delete_account(username):
    wb = openpyxl.load_workbook("data.xlsx")
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        if username == sheet.cell(row=i, column=1).value:
            sheet.cell(row=i, column=1).value = ""
            sheet.cell(row=i, column=2).value = ""
            wb.save("data.xlsx")
            flash("Account Deleted!")
    return redirect(url_for("accounts"))

@app.route("/new_trade", methods=["POST", "GET"])
def new_trade():
    if request.method == "POST":
        dic_qty = {}
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value}

        call_sell = request.form["call_sell"]
        call_buy = request.form["call_buy"]
        put_sell = request.form["put_sell"]
        put_buy = request.form["put_buy"]
        expiry_sell = request.form["expiry_sell"]
        expiry_hedge = request.form["expiry_hedge"]

        for user in usernames:
            dic_qty[user] = request.form[user]
        try:
            for key, value in dic_qty.items():
                if value != '0':
                    t1 = threading.Thread(target=take_new_trade, args=(key, usernames[key], call_sell, call_buy, put_sell, put_buy, value, expiry_sell, expiry_hedge))
                    t1.start()
            flash("Trades Taken!")
        except Exception as e:
            print(e)
            flash("Some error occurred!")
        return redirect(url_for("new_trade"))
    else:
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]
        return render_template("new_trade.html", usernames=usernames)

@app.route("/shifting",methods=["POST","GET"])
def shifting():
    if request.method == "POST":
        dic_qty = {}
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = {}
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames[cell] = sheet.cell(row=i, column=2).value

        previous_call_sold = request.form["current_call_sold"]
        previous_call_hedge = request.form["current_call_hedge"]
        new_call_sell = request.form["new_call_sell"]
        new_call_hedge = request.form["new_call_hedge"]

        previous_put_sold = request.form["current_put_sold"]
        previous_put_hedge = request.form["current_put_hedge"]
        new_put_sell = request.form["new_put_sell"]
        new_put_hedge = request.form["new_put_hedge"]

        previous_expiry_sell = request.form["previous_expiry_sell"]
        current_expiry_sell = request.form["current_expiry_sell"]

        previous_expiry_hedge = request.form["previous_expiry_hedge"]
        current_expiry_hedge = request.form["current_expiry_hedge"]

        for key,value in usernames.items():
            dic_qty[key] = request.form[key]

        for key,value in dic_qty.items():
            if value!='0':
                t1 = threading.Thread(target=shift, args=(key,usernames[key],previous_call_sold, new_call_sell,previous_call_hedge,new_call_hedge, previous_put_sold, new_put_sell,previous_put_hedge,new_put_hedge,previous_expiry_sell,current_expiry_sell,previous_expiry_hedge,current_expiry_hedge,value,))
                t1.start()
        flash("Shifting Done!")
        return redirect(url_for("shifting"))

    else:
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        usernames = []
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1).value
            if cell != None:
                usernames.append(cell)
        return render_template("shifting.html",usernames=usernames)
        
@app.route('/downloads')
def download_home():
    return render_template('index.html')

@app.route('/download', methods=['POST'])
def download():
    symbol = request.form['symbol']
    backtest_years = int(request.form['backtest_years'])
    timeframe = request.form['timeframe']

    output_filename = f"{symbol}_backtest.xlsx"
    output_filepath = os.path.join('downloads', output_filename)

    try:
        backtest_to_excel(symbol, backtest_years, timeframe, output_filepath)
        return send_file(output_filepath, as_attachment=True)
    except Exception as e:
        return f"Error: {str(e)}"

def take_new_trade(username, api_key, call_sell, call_buy, put_sell, put_buy, qty,expiry_sell,expiry_hedge):
    qty = int(qty)
    alice = Aliceblue(user_id=username, api_key=api_key)
    aliceblue_Res = alice.get_session_id()
    print(aliceblue_Res)
    alice.get_contract_master("NFO")
    a = int(qty / 1800)
    for i in range(0, a):
        PlaceBuyOrder(alice,1800,True,call_buy,expiry_hedge)
        PlaceSellOrder(alice, 1800, True, call_sell, expiry_sell)
        PlaceBuyOrder(alice,1800,False,put_buy,expiry_hedge)
        PlaceSellOrder(alice, 1800, False, put_sell, expiry_sell)

    PlaceBuyOrder(alice, qty-(1800*a), True, call_buy, expiry_hedge)
    PlaceSellOrder(alice, qty-(1800*a), True, call_sell, expiry_sell)
    PlaceBuyOrder(alice, qty-(1800*a), False, put_buy, expiry_hedge)
    PlaceSellOrder(alice, qty-(1800*a), False, put_sell, expiry_sell)

def shift(username, api_key, previous_call_sold, new_call_sell,previous_call_hedge,new_call_hedge, previous_put_sold, new_put_sell,previous_put_hedge,new_put_hedge,previous_expiry_sell,current_expiry_sell,previous_expiry_hedge,current_expiry_hedge,qty):
    qty = int(qty)
    alice = Aliceblue(user_id=username, api_key=api_key)
    aliceblue_Res = alice.get_session_id()
    print(aliceblue_Res)
    alice.get_contract_master("NFO")

    #take new hedge

    a = int(qty / 1800)
    for i in range(0, a):
        if int(previous_put_hedge) != 1 and int(new_put_hedge) != 1:
            PlaceBuyOrder(alice, 1800, False, new_put_hedge, current_expiry_hedge)

        if int(previous_call_hedge) != 1 and int(new_call_hedge) != 1:
            PlaceBuyOrder(alice, 1800, True, new_call_hedge, current_expiry_hedge)
            
    if int(previous_put_hedge) != 1 and int(new_put_hedge) != 1:
        PlaceBuyOrder(alice, qty - (1800*a), False, new_put_hedge, current_expiry_hedge)

    if int(previous_call_hedge) != 1 and int(new_call_hedge) != 1:
        PlaceBuyOrder(alice, qty - (1800*a), True, new_call_hedge, current_expiry_hedge)

    #squareoff currently sold put
    
    a = int(qty / 1800)
    for i in range(0,a):
        if int(previous_put_sold)!=1 and int(new_put_sell)!=1:
            PlaceBuyOrder(alice,1800,False,previous_put_sold,previous_expiry_sell)

        if int(previous_call_sold)!=1 and int(new_call_sell)!=1:
            PlaceBuyOrder(alice,1800,True,previous_call_sold,previous_expiry_sell)
            
        if int(previous_put_sold)!=1 and int(new_put_sell)!=1:
            PlaceSellOrder(alice, 1800, False,new_put_sell,current_expiry_sell)

        if int(previous_call_sold)!=1 and int(new_call_sell)!=1:
            PlaceSellOrder(alice, 1800, True,new_call_sell,current_expiry_sell)

    if int(previous_put_sold)!=1 and int(new_put_sell)!=1:
        PlaceBuyOrder(alice,qty - (1800*a),False,previous_put_sold,previous_expiry_sell)

    if int(previous_call_sold)!=1 and int(new_call_sell)!=1:
        PlaceBuyOrder(alice,qty - (1800*a),True,previous_call_sold,previous_expiry_sell)
            
    if int(previous_put_sold)!=1 and int(new_put_sell)!=1:
        PlaceSellOrder(alice, qty - (1800*a), False,new_put_sell,current_expiry_sell)

    if int(previous_call_sold)!=1 and int(new_call_sell)!=1:
        PlaceSellOrder(alice, qty - (1800*a), True,new_call_sell,current_expiry_sell)



    #square off old hedge
    
    a = int(qty / 1800)
    for i in range(0, a):
        if int(previous_put_hedge) != 1 and int(new_put_hedge) != 1:
            PlaceSellOrder(alice, 1800, False, previous_put_hedge, previous_expiry_hedge)

        if int(previous_call_hedge) != 1 and int(new_call_hedge) != 1:
            PlaceSellOrder(alice, 1800, True, previous_call_hedge, previous_expiry_hedge)
            
    if int(previous_put_hedge) != 1 and int(new_put_hedge) != 1:
        PlaceSellOrder(alice, qty - (1800*a), False, previous_put_hedge, previous_expiry_hedge)

    if int(previous_call_hedge) != 1 and int(new_call_hedge) != 1:
        PlaceSellOrder(alice, qty - (1800*a), True, previous_call_hedge, previous_expiry_hedge)
        
        

def PlaceBuyOrder(alice, qty, call,strike,expiry):
    alice.get_contract_master("NFO")
    if int(strike)!=0 or int(strike)!=1:
        res_2 = alice.place_order(transaction_type=TransactionType.Buy,
                                instrument=alice.get_instrument_for_fno(exch="NFO",symbol='NIFTY', expiry_date=expiry, is_fut=False,strike=int(strike), is_CE=call),
                                quantity=qty,
                                order_type=OrderType.Market,
                                product_type=ProductType.Normal,
                                price=0.0,
                                trigger_price=None,
                                stop_loss=None,
                                square_off=None,
                                trailing_sl=None,
                                is_amo=False,
                                order_tag='order1')
    t.sleep(2)

def PlaceSellOrder(alice, qty, call,strike,expiry):
    alice.get_contract_master("NFO")
    if int(strike)!=0 or int(strike)!=1:
        res_2 = alice.place_order(transaction_type=TransactionType.Sell,
                                instrument=alice.get_instrument_for_fno(exch="NFO",symbol='NIFTY', expiry_date=expiry, is_fut=False,strike=int(strike), is_CE=call),
                                quantity=qty,
                                order_type=OrderType.Market,
                                product_type=ProductType.Normal,
                                price=0.0,
                                trigger_price=None,
                                stop_loss=None,
                                square_off=None,
                                trailing_sl=None,
                                is_amo=False,
                                order_tag='order1')
    t.sleep(2)

if __name__ == "__main__":
    app.run(debug=True)
